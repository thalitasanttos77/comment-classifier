import os
import argparse
import time
import numpy as np

from embedder import embed_text, tokenize
from model_io import load_resources, predict_proba_array

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as e:
    raise SystemExit("openpyxl não está instalado. Instale com:\n  pip install openpyxl") from e

# Heurística opcional
try:
    from heuristics import adjust_probability_with_heuristics
except Exception:
    adjust_probability_with_heuristics = None

def autosize_columns(ws):
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

def safe_save_workbook(wb, path: str) -> str:
    try:
        wb.save(path)
        return os.path.abspath(path)
    except PermissionError:
        root, ext = os.path.splitext(path)
        ts = time.strftime("%Y%m%d-%H%M%S")
        alt = f"{root}_{ts}{ext or '.xlsx'}"
        wb.save(alt)
        return os.path.abspath(alt)

def main():
    parser = argparse.ArgumentParser(
        description="Classifica um arquivo (1 comentário por linha) e gera Excel (.xlsx): texto | classificacao | probabilidade."
    )
    parser.add_argument("--model_dir", type=str, default="models")
    parser.add_argument("--input_txt", type=str, required=True)
    parser.add_argument("--output_xlsx", type=str, default="predicoes.xlsx")
    parser.add_argument("--strip_accents", action="store_true")
    parser.add_argument("--sheet_name", type=str, default="Classificacoes")
    parser.add_argument("--threshold", type=float, default=None, help="Se não passar, usa o salvo no modelo")
    parser.add_argument("--use_heuristics", action="store_true", help="Aplicar regras simples (negação/léxico)")
    args = parser.parse_args()

    model, scaler, word_map, saved_th, model_type = load_resources(args.model_dir, args.strip_accents)
    decision_th = float(args.threshold) if args.threshold is not None else saved_th

    with open(args.input_txt, "r", encoding="utf-8") as f:
        texts = [line.strip() for line in f if line.strip()]

    wb = Workbook()
    ws = wb.active
    ws.title = args.sheet_name

    headers = ["texto", "classificacao", "probabilidade"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    for i in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=i)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    X = []
    toks_list = []
    for text in texts:
        toks = tokenize(text, lowercase=True, strip_accents=args.strip_accents)
        toks_list.append(toks)
        X.append(embed_text(text, word_map, lowercase=True, strip_accents=args.strip_accents))
    X = np.stack(X, axis=0)
    X_s = scaler.transform(X)

    proba = predict_proba_array(model, X_s, model_type).astype(float)
    if args.use_heuristics and adjust_probability_with_heuristics is not None:
        proba = np.array([adjust_probability_with_heuristics(float(p), toks) for p, toks in zip(proba, toks_list)], dtype=float)
    preds = (proba >= decision_th).astype(int)

    for text, pred, p in zip(texts, preds, proba):
        label = "positivo" if pred == 1 else "negativo"
        conf = p if pred == 1 else (1.0 - p)

        r = ws.max_row + 1
        ws.cell(row=r, column=1, value=text)
        ws.cell(row=r, column=2, value=label)
        c3 = ws.cell(row=r, column=3, value=float(conf))
        c3.number_format = "0%"
        c3.alignment = center

    autosize_columns(ws)

    out = args.output_xlsx
    if not out.lower().endswith(".xlsx"):
        out += ".xlsx"
    abs_path = safe_save_workbook(wb, out)
    print(f"Planilha Excel salva em: {abs_path}")

if __name__ == "__main__":
    main()